"""
Excel 詳細計算表匯出模組（繁體中文版）
使用 openpyxl 產生含公式的 .xlsx 檔案
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

HEADER_FILL = PatternFill('solid', fgColor='1A5276')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10, name='Arial')
ALT_FILL = PatternFill('solid', fgColor='EAF2F8')
TITLE_FONT = Font(bold=True, size=14, color='1A5276', name='Arial')
SUBTITLE_FONT = Font(bold=True, size=11, color='2C3E50', name='Arial')
NORMAL_FONT = Font(size=10, name='Arial')
FORMULA_FONT = Font(size=10, color='0000FF', name='Arial')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_alt=False):
    cell = ws.cell(row=row, column=col)
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = CENTER
    if is_alt:
        cell.fill = ALT_FILL
    return cell


def generate_xlsx(results, params, sensitivity_data=None):
    wb = Workbook()

    # ===== Sheet 1: 輸入參數 =====
    ws1 = wb.active
    ws1.title = '輸入參數'
    ws1.sheet_properties.tabColor = '1A5276'

    ws1.merge_cells('A1:C1')
    ws1['A1'] = '冷卻水塔 - 輸入參數'
    ws1['A1'].font = TITLE_FONT

    ws1.merge_cells('A2:C2')
    ws1['A2'] = f"專案：{params.get('project_name', '-')}  |  計算人員：{params.get('engineer', '-')}  |  日期：{params.get('date', datetime.now().strftime('%Y-%m-%d'))}"
    ws1['A2'].font = Font(size=9, color='666666', name='Arial')

    row = 4
    for col, h in enumerate(['參數', '數值', '單位'], 1):
        ws1.cell(row=row, column=col, value=h)
    style_header_row(ws1, row, 3)

    tower_label = "逆流式" if params.get('tower_type') == 'counter' else "橫流式"
    season_label = "夏季 (C=1.0)" if params.get('season') == 'summer' else "冬季 (C=0.8)"

    param_rows = [
        ('循環水量 (Q)', params['Q'], 'm³/h'),
        ('進水溫度 (T_in)', params['T_in'], '°C'),
        ('出水溫度 (T_out)', params['T_out'], '°C'),
        ('溫差 (ΔT)', params['T_in'] - params['T_out'], '°C'),
        ('乾球溫度 (Tdb)', params['Tdb'], '°C'),
        ('濕球溫度 (Twb)', params['Twb'], '°C'),
        ('空氣流量 (G)', params['G'], 'm³/h'),
        ('填料特性係數 (KaV/L)', params['KaV_L'], '-'),
        ('濃縮倍數 (COC)', params['COC'], '-'),
        ('水塔類型', tower_label, '-'),
        ('季節', season_label, '-'),
        ('大氣壓力', params.get('P', 101.325), 'kPa'),
        ('水的密度', 1000, 'kg/m³'),
        ('水的比熱', 4.186, 'kJ/(kg·°C)'),
        ('空氣密度（近似）', 1.2, 'kg/m³'),
    ]

    for i, (name, val, unit) in enumerate(param_rows):
        r = row + 1 + i
        is_alt = i % 2 == 1
        ws1.cell(row=r, column=1, value=name)
        style_data_cell(ws1, r, 1, is_alt).alignment = LEFT
        ws1.cell(row=r, column=2, value=val)
        style_data_cell(ws1, r, 2, is_alt)
        ws1.cell(row=r, column=3, value=unit)
        style_data_cell(ws1, r, 3, is_alt)

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18

    # ===== Sheet 2: 計算過程 =====
    ws2 = wb.create_sheet('計算過程')
    ws2.sheet_properties.tabColor = '27AE60'

    ws2.merge_cells('A1:D1')
    ws2['A1'] = '逐步計算過程'
    ws2['A1'].font = TITLE_FONT

    row = 3
    for col, h in enumerate(['步驟', '公式', '代入數值', '計算結果'], 1):
        ws2.cell(row=row, column=col, value=h)
    style_header_row(ws2, row, 4)

    Q = params['Q']
    T_in = params['T_in']
    T_out = params['T_out']
    dT = T_in - T_out
    G = params['G']
    COC = params['COC']
    md = results['merkel_details']
    C = 1.0 if params.get('season') == 'summer' else 0.8
    splash_rate = 0.001 if params.get('tower_type') == 'counter' else 0.003

    steps = [
        ('1. 溫差計算', 'ΔT = T_in - T_out',
         f'{T_in} - {T_out}', f'{dT:.1f} °C'),
        ('2. 蒸發係數 (C)', f'C = {"1.0（夏季）" if C==1.0 else "0.8（冬季）"}',
         '-', f'{C}'),
        ('3. 簡易公式蒸發量', 'E = C × ΔT × Q / 600',
         f'{C} × {dT:.1f} × {Q:.1f} / 600', f'{results["E_simple"]:.4f} m³/h'),
        ('', '', '', ''),
        ('4. 水的質量流率 (L)', 'L = Q × ρw / 3600',
         f'{Q:.1f} × 1000 / 3600', f'{Q*1000/3600:.2f} kg/s'),
        ('5. 空氣質量流率 (G_m)', 'G_m = G × ρa / 3600',
         f'{G:.1f} × 1.2 / 3600', f'{G*1.2/3600:.2f} kg/s'),
        ('6. 氣水比 (L/G)', 'L/G = L / G_m',
         f'{Q*1000/3600:.2f} / {G*1.2/3600:.2f}', f'{md["LG_ratio"]:.3f}'),
        ('7. Merkel 數', 'NTU = ∫(Cp·dT)/(h_s - h_a)',
         '數值積分（辛普森法）', f'{md["merkel_number"]:.3f}'),
        ('8. Merkel 法蒸發量', '由焓平衡計算',
         'G_m × ΔW', f'{results["E_merkel"]:.4f} m³/h'),
        ('', '', '', ''),
        ('9. 飛濺損失', f'損失率 = {splash_rate*100:.1f}% × Q',
         f'{splash_rate} × {Q:.1f}', f'{results["E_splash"]:.4f} m³/h'),
        ('10. 排污損失', '排污 = E / (COC - 1)',
         f'{results["E_merkel"]:.4f} / ({COC:.1f} - 1)', f'{results["E_blowdown"]:.4f} m³/h'),
        ('11. 總補給水量', '合計 = 蒸發 + 飛濺 + 排污',
         f'{results["E_merkel"]:.4f} + {results["E_splash"]:.4f} + {results["E_blowdown"]:.4f}',
         f'{results["E_total"]:.4f} m³/h'),
        ('', '', '', ''),
        ('12. 逼近溫度', '逼近 = T_out - T_wb',
         f'{T_out:.1f} - {md["T_approach"] + T_out - T_out:.1f}',
         f'{md["T_approach"]:.1f} °C'),
        ('13. 冷卻範圍', '範圍 = T_in - T_out',
         f'{T_in:.1f} - {T_out:.1f}', f'{md["T_range"]:.1f} °C'),
        ('14. 冷卻效率', 'η = 範圍 / (範圍 + 逼近)',
         f'{md["T_range"]:.1f} / ({md["T_range"]:.1f} + {md["T_approach"]:.1f})',
         f'{md["efficiency"]:.1f} %'),
    ]

    for i, (step, formula, sub, result) in enumerate(steps):
        r = row + 1 + i
        is_alt = i % 2 == 1
        ws2.cell(row=r, column=1, value=step)
        style_data_cell(ws2, r, 1, is_alt).alignment = LEFT
        ws2.cell(row=r, column=2, value=formula)
        style_data_cell(ws2, r, 2, is_alt).alignment = LEFT
        if formula:
            ws2.cell(row=r, column=2).font = FORMULA_FONT
        ws2.cell(row=r, column=3, value=sub)
        style_data_cell(ws2, r, 3, is_alt).alignment = LEFT
        ws2.cell(row=r, column=4, value=result)
        style_data_cell(ws2, r, 4, is_alt)

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 22

    # ===== Sheet 3: 結果摘要（含 Excel 公式）=====
    ws3 = wb.create_sheet('結果摘要')
    ws3.sheet_properties.tabColor = 'E74C3C'

    ws3.merge_cells('A1:C1')
    ws3['A1'] = '結果摘要（含即時公式）'
    ws3['A1'].font = TITLE_FONT

    ws3['A3'] = '可編輯輸入值'
    ws3['A3'].font = SUBTITLE_FONT

    row = 4
    for col, h in enumerate(['參數', '數值', '單位'], 1):
        ws3.cell(row=row, column=col, value=h)
    style_header_row(ws3, row, 3)

    input_cells = {}
    editable_inputs = [
        ('Q', '循環水量', Q, 'm³/h'),
        ('T_in', '進水溫度', T_in, '°C'),
        ('T_out', '出水溫度', T_out, '°C'),
        ('C', '蒸發係數', C, '-'),
        ('COC', '濃縮倍數', COC, '-'),
        ('splash_rate', '飛濺損失率', splash_rate, '-'),
    ]

    for i, (key, name, val, unit) in enumerate(editable_inputs):
        r = row + 1 + i
        ws3.cell(row=r, column=1, value=name)
        style_data_cell(ws3, r, 1).alignment = LEFT
        cell_val = ws3.cell(row=r, column=2, value=val)
        cell_val.font = Font(size=10, color='0000FF', name='Arial', bold=True)
        cell_val.border = THIN_BORDER
        cell_val.alignment = CENTER
        ws3.cell(row=r, column=3, value=unit)
        style_data_cell(ws3, r, 3)
        input_cells[key] = f'B{r}'

    calc_start = row + len(editable_inputs) + 2
    ws3.cell(row=calc_start, column=1, value='計算結果（自動更新）')
    ws3[f'A{calc_start}'].font = SUBTITLE_FONT

    calc_row = calc_start + 1
    for col, h in enumerate(['項目', '公式說明', '計算結果', '單位'], 1):
        ws3.cell(row=calc_row, column=col, value=h)
    style_header_row(ws3, calc_row, 4)

    q_ref = input_cells['Q']
    tin_ref = input_cells['T_in']
    tout_ref = input_cells['T_out']
    c_ref = input_cells['C']
    coc_ref = input_cells['COC']
    sr_ref = input_cells['splash_rate']

    formulas = [
        ('ΔT（溫差）', f'={tin_ref}-{tout_ref}', '°C'),
        ('簡易公式蒸發量', f'={c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600', 'm³/h'),
        ('飛濺損失', f'={sr_ref}*{q_ref}', 'm³/h'),
        ('排污損失（估算）', f'={c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600/({coc_ref}-1)', 'm³/h'),
        ('總補給水量（簡易）', f'={c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600+{sr_ref}*{q_ref}+{c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600/({coc_ref}-1)', 'm³/h'),
        ('每日補給水量', f'=({c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600+{sr_ref}*{q_ref}+{c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600/({coc_ref}-1))*24', 'm³/日'),
        ('每月補給水量', f'=({c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600+{sr_ref}*{q_ref}+{c_ref}*({tin_ref}-{tout_ref})*{q_ref}/600/({coc_ref}-1))*24*30', 'm³/月'),
    ]

    for i, (name, formula, unit) in enumerate(formulas):
        r = calc_row + 1 + i
        is_alt = i % 2 == 1
        ws3.cell(row=r, column=1, value=name)
        style_data_cell(ws3, r, 1, is_alt).alignment = LEFT
        ws3.cell(row=r, column=2, value=formula.replace('=', ''))
        style_data_cell(ws3, r, 2, is_alt).alignment = LEFT
        ws3.cell(row=r, column=2).font = Font(size=9, color='888888', name='Arial')
        ws3.cell(row=r, column=3, value=formula)
        style_data_cell(ws3, r, 3, is_alt)
        ws3.cell(row=r, column=3).font = FORMULA_FONT
        ws3.cell(row=r, column=4, value=unit)
        style_data_cell(ws3, r, 4, is_alt)

    total_row = calc_row + 5
    ws3.conditional_formatting.add(
        f'C{calc_row+1}:C{total_row}',
        CellIsRule(operator='greaterThan', formula=['100'],
                   fill=PatternFill('solid', fgColor='FFCCCC'))
    )

    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 45
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 15

    # ===== Sheet 4: 敏感度分析 =====
    ws4 = wb.create_sheet('敏感度分析')
    ws4.sheet_properties.tabColor = 'F39C12'

    ws4.merge_cells('A1:H1')
    ws4['A1'] = '敏感度分析矩陣'
    ws4['A1'].font = TITLE_FONT

    # ΔT 敏感度
    ws4['A3'] = '溫差 (ΔT) 對蒸發量的影響'
    ws4['A3'].font = SUBTITLE_FONT

    dt_values = [2, 3, 4, 5, 6, 7, 8, 10, 12, 15]
    row = 4
    ws4.cell(row=row, column=1, value='ΔT (°C)')
    ws4.cell(row=row, column=2, value='簡易蒸發量 (m³/h)')
    ws4.cell(row=row, column=3, value='總補給水量 (m³/h)')
    style_header_row(ws4, row, 3)

    for i, dt_val in enumerate(dt_values):
        r = row + 1 + i
        ws4.cell(row=r, column=1, value=dt_val)
        style_data_cell(ws4, r, 1, i % 2 == 1)
        ws4.cell(row=r, column=2, value=f"='結果摘要'!{c_ref}*A{r}*'結果摘要'!{q_ref}/600")
        style_data_cell(ws4, r, 2, i % 2 == 1).font = FORMULA_FONT
        ws4.cell(row=r, column=3, value=f"='結果摘要'!{c_ref}*A{r}*'結果摘要'!{q_ref}/600+'結果摘要'!{sr_ref}*'結果摘要'!{q_ref}+'結果摘要'!{c_ref}*A{r}*'結果摘要'!{q_ref}/600/('結果摘要'!{coc_ref}-1)")
        style_data_cell(ws4, r, 3, i % 2 == 1).font = FORMULA_FONT

    # COC 敏感度
    coc_start = row + len(dt_values) + 3
    ws4.cell(row=coc_start, column=1, value='濃縮倍數 (COC) 對排污量的影響')
    ws4[f'A{coc_start}'].font = SUBTITLE_FONT

    coc_values = [1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 5.0, 6.0, 8.0, 10.0]
    row_coc = coc_start + 1
    ws4.cell(row=row_coc, column=1, value='COC')
    ws4.cell(row=row_coc, column=2, value='排污量 (m³/h)')
    ws4.cell(row=row_coc, column=3, value='總補給水量 (m³/h)')
    ws4.cell(row=row_coc, column=4, value='相對 COC=2 節水率 (%)')
    style_header_row(ws4, row_coc, 4)

    for i, coc_val in enumerate(coc_values):
        r = row_coc + 1 + i
        ws4.cell(row=r, column=1, value=coc_val)
        style_data_cell(ws4, r, 1, i % 2 == 1)
        ws4.cell(row=r, column=2, value=f"='結果摘要'!{c_ref}*('結果摘要'!{tin_ref}-'結果摘要'!{tout_ref})*'結果摘要'!{q_ref}/600/(A{r}-1)")
        style_data_cell(ws4, r, 2, i % 2 == 1).font = FORMULA_FONT
        ws4.cell(row=r, column=3, value=f"='結果摘要'!{c_ref}*('結果摘要'!{tin_ref}-'結果摘要'!{tout_ref})*'結果摘要'!{q_ref}/600+'結果摘要'!{sr_ref}*'結果摘要'!{q_ref}+'結果摘要'!{c_ref}*('結果摘要'!{tin_ref}-'結果摘要'!{tout_ref})*'結果摘要'!{q_ref}/600/(A{r}-1)")
        style_data_cell(ws4, r, 3, i % 2 == 1).font = FORMULA_FONT
        coc2_row = row_coc + 1 + 1
        ws4.cell(row=r, column=4, value=f'=IF(C{coc2_row}=0,0,(1-C{r}/C{coc2_row})*100)')
        style_data_cell(ws4, r, 4, i % 2 == 1).font = FORMULA_FONT

    # 流量敏感度
    q_start = row_coc + len(coc_values) + 3
    ws4.cell(row=q_start, column=1, value='循環水量 (Q) 對蒸發量的影響')
    ws4[f'A{q_start}'].font = SUBTITLE_FONT

    q_values = [100, 200, 300, 500, 750, 1000, 1500, 2000]
    row_q = q_start + 1
    ws4.cell(row=row_q, column=1, value='Q (m³/h)')
    ws4.cell(row=row_q, column=2, value='蒸發量 (m³/h)')
    ws4.cell(row=row_q, column=3, value='總補給水量 (m³/h)')
    style_header_row(ws4, row_q, 3)

    for i, q_val in enumerate(q_values):
        r = row_q + 1 + i
        ws4.cell(row=r, column=1, value=q_val)
        style_data_cell(ws4, r, 1, i % 2 == 1)
        ws4.cell(row=r, column=2, value=f"='結果摘要'!{c_ref}*('結果摘要'!{tin_ref}-'結果摘要'!{tout_ref})*A{r}/600")
        style_data_cell(ws4, r, 2, i % 2 == 1).font = FORMULA_FONT
        ws4.cell(row=r, column=3, value=f"='結果摘要'!{c_ref}*('結果摘要'!{tin_ref}-'結果摘要'!{tout_ref})*A{r}/600+'結果摘要'!{sr_ref}*A{r}+'結果摘要'!{c_ref}*('結果摘要'!{tin_ref}-'結果摘要'!{tout_ref})*A{r}/600/('結果摘要'!{coc_ref}-1)")
        style_data_cell(ws4, r, 3, i % 2 == 1).font = FORMULA_FONT

    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 22
    ws4.column_dimensions['D'].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
